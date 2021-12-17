from os import environ
import boto3
import json
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment

region_mapping = {"ap-southeast-1": 'Singapore', "ap-east-1": 'Hong Kong'}

# Route table variables
route_header_row = 2
vpc_row = 2
route_table_row = 2
route_row = 2

# Transit Gateway variables
transit_gateway_header_row = 2
transit_gateway_center_row = 2
transit_gateway_row = 2


def get_vpc_data(role, accoutnId, region, process_type):
    ws = wb.worksheets[sheet_list.index(process_type) + 1]
    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])
    response = client.describe_vpcs()
    vpc_name = ''

    for vpc in response['Vpcs']:
        if 'Tags' in vpc:
            for tag in vpc['Tags']:
                if tag['Key'] == 'Name':
                    vpc_name = tag['Value']
        ws.append([vpc_name, vpc['CidrBlock'],
                   vpc['VpcId'], accoutnId, region_mapping[region]])


def get_subnet_data(role, region):
    ws = wb.create_sheet("Subnet")
    column_list = ['Subnet Name', 'CIDR', 'VPC ID', 'AZ', 'Subnet ID']
    ws.append(column_list)

    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])
    response = client.describe_subnets()
    for subnet in response['Subnets']:
        vpc_name = subnet['Tags']
        for tag in vpc_name:
            if tag['Key'] == 'Name':
                ws.append([tag['Value'], subnet['CidrBlock'], subnet['VpcId'],
                          subnet['AvailabilityZone'], subnet['SubnetId']])


def get_route_table_data(role, accountId, region, process_type):

    global route_header_row
    global vpc_row
    global route_table_row
    global route_row

    def get_target_id(route):
        if 'GatewayId' in route:
            return route['GatewayId']
        elif 'InstanceId' in route:
            return route['InstanceId']
        elif 'NetworkInterfaceId' in route:
            return route['NetworkInterfaceId']
        elif 'VpcPeeringConnectionId' in route:
            return route['VpcPeeringConnectionId']
        elif 'NatGatewayId' in route:
            return route['NatGatewayId']
        elif 'TransitGatewayId' in route:
            return route['TransitGatewayId']
        elif 'VpcEndpointId' in route:
            return route['VpcEndpointId']

    ws = wb.worksheets[sheet_list.index(process_type) + 1]
    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])
    response = client.describe_route_tables()

    output = {'total': 0}
    name = ''
    for route_table in response['RouteTables']:
        for tag in route_table['Tags']:
            if tag['Key'] == 'Name':
                name = tag['Value']
        if route_table['VpcId'] in output:
            output[route_table['VpcId']][route_table['RouteTableId']
                                         ] = {
                                             'name': name,
                                             'routes': route_table['Routes'],
                                             'total': len(route_table['Routes'])
            }
            output[route_table['VpcId']
                   ]['total'] += len(route_table['Routes'])
        else:
            output[route_table['VpcId']] = {}
            output[route_table['VpcId']][route_table['RouteTableId']
                                         ] = {
                                             'name': name,
                                             'routes': route_table['Routes'],
                                             'total': len(route_table['Routes'])
            }
            output[route_table['VpcId']
                   ]['total'] = len(route_table['Routes'])
        output['total'] += len(route_table['Routes'])
    # print(json.dumps(output, indent=4))

    # Write to Excel
    if output['total'] > 0:
        # Region Column
        ws.merge_cells(start_row=route_header_row, start_column=1,
                       end_row=route_header_row + output['total'] - 1, end_column=1)
        ws.cell(row=route_header_row,
                column=1).value = region_mapping[region]
        # Environment Column
        ws.merge_cells(start_row=route_header_row, start_column=2,
                       end_row=route_header_row + output['total'] - 1, end_column=2)
        ws.cell(row=route_header_row, column=2).value = environment
        # Account Column
        ws.merge_cells(start_row=route_header_row, start_column=3,
                       end_row=route_header_row + output['total'] - 1, end_column=3)
        ws.cell(row=route_header_row, column=3).value = accountId
        route_header_row += output['total']

        for vpc_id in output:
            if vpc_id == 'total':
                continue
            ws.merge_cells(start_row=vpc_row, start_column=4,
                           end_row=vpc_row+output[vpc_id]['total'] - 1, end_column=4)
            ws.cell(row=vpc_row, column=4).value = vpc_id
            vpc_row += output[vpc_id]['total']
            for route_table in output[vpc_id]:
                if route_table == 'total':
                    continue
                ws.merge_cells(start_row=route_table_row, start_column=5,
                               end_row=route_table_row+output[vpc_id][route_table]['total'] - 1, end_column=5)
                ws.cell(row=route_table_row,
                        column=5).value = output[vpc_id][route_table]['name']
                ws.merge_cells(start_row=route_table_row, start_column=6,
                               end_row=route_table_row+output[vpc_id][route_table]['total'] - 1, end_column=6)
                ws.cell(row=route_table_row,
                        column=6).value = route_table
                route_table_row += output[vpc_id][route_table]['total']
                for route in output[vpc_id][route_table]['routes']:
                    if 'DestinationCidrBlock' in route:
                        ws.cell(row=route_row,
                                column=7).value = route['DestinationCidrBlock']
                    elif 'DestinationPrefixListId' in route:
                        ws.cell(row=route_row,
                                column=7).value = route['DestinationPrefixListId']
                    ws.cell(row=route_row, column=8).value = get_target_id(route)
                    route_row += 1


def get_igw_data(role, accountId, region):
    ws = wb.create_sheet('Internet Gateway')
    column_list = ["Region", "Environment", "Account", "Type",
                   'IGW Name', 'IGW ID', "VPC ID"]
    ws.append(column_list)

    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])
    response = client.describe_internet_gateways()
    for igw in response['InternetGateways']:
        igw_name = igw['Tags']
        for tag in igw_name:
            if tag['Key'] == 'Name':
                ws.append([region_mapping[region], environment, accountId, 'igw', tag['Value'], igw['InternetGatewayId'],
                           igw['Attachments'][0]['VpcId']])


def get_endpoint_data(role, region):
    ws = wb.create_sheet('VPC Endpoint')
    column_list = ['VPC Endpoint Name', 'VPC Endpoint ID', 'VPC ID']
    ws.append(column_list)

    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])
    response = client.describe_vpc_endpoints()
    for endpoint in response['VpcEndpoints']:
        endpoint_name = endpoint['Tags']
        for tag in endpoint_name:
            if tag['Key'] == 'Name':
                ws.append([tag['Value'], endpoint['VpcEndpointId'],
                          endpoint['VpcId']])


def get_endpoint_service_data(role, region):
    ws = wb.create_sheet('VPC Endpoint Service')
    column_list = ['VPC Endpoint Service Name', 'VPC Endpoint Service ID',
                   'VPC Endpoint Service Type', 'VPC Endpoint Service AZ']
    ws.append(column_list)

    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])
    response = client.describe_vpc_endpoint_service_configurations()
    for endpoint_service in response['ServiceConfigurations']:
        print(json.dumps(endpoint_service, indent=4))
        endpoint_service_type = endpoint_service['ServiceType'][0]['ServiceType']
        endpoint_service_id = endpoint_service['ServiceId']
        endpoint_services_available = ','.join(
            map(str, endpoint_service['AvailabilityZones']))
        for tag in endpoint_service['Tags']:
            if tag['Key'] == 'Name':
                ws.append([tag['Value'], endpoint_service_id, endpoint_service_type,
                          endpoint_services_available])


def get_nat_data(role, region):
    ws = wb.create_sheet('NAT Gateway')
    column_list = ['NAT Gateway Name', 'NAT Gateway ID', 'VPC ID']
    ws.append(column_list)

    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])
    response = client.describe_nat_gateways()
    for nat in response['NatGateways']:
        nat_name = nat['Tags']
        for tag in nat_name:
            if tag['Key'] == 'Name':
                ws.append([tag['Value'], nat['NatGatewayId'], nat['VpcId']])


def get_sg_data(role, region):
    ws = wb.create_sheet('Security Group')
    column_list = ['Security Group Name',
                   'Security Group ID', 'VPC ID', "Ingress", "Egress"]
    ws.append(column_list)

    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])
    response = client.describe_security_groups()
    for sg in response['SecurityGroups']:
        ingress = ''
        egress = ''
        if len(sg['IpPermissions']) > 0:
            for ing in sg['IpPermissions']:
                if ing['IpProtocol'] == '-1':
                    ingress += (
                        '{}:{}\n'.format(ing['IpProtocol'], ing['IpRanges'][0]['CidrIp']))
                else:
                    ingress += (
                        '{}:{}:{}\n'.format(ing['IpProtocol'], ing['IpRanges'][0]['CidrIp'], ing['FromPort']))
        # print(ingress)
        for eg in sg['IpPermissionsEgress']:
            # print(json.dumps(eg, indent=4))
            if eg['IpProtocol'] == '-1':
                egress += (
                    '{}:{}\n'.format(eg['IpProtocol'], eg['IpRanges'][0]['CidrIp']))
            else:
                egress += (
                    '{}:{}:{}\n'.format(eg['IpProtocol'], eg['IpRanges'][0]['CidrIp'], eg['FromPort']))
        ws.append([sg['GroupName'], sg['GroupId'],
                  sg['VpcId'], ingress, egress])


def get_vpn_data(role, accountId,  region):
    ws = wb.create_sheet('VPN Connection')
    column_list = ['Region', 'Environment', 'Type', 'Ou',
                   'Project Name', 'Location', 'Environment', 'AccountId', 'No',
                   'VPN Connection Name', 'Description', 'VPC ID', 'VPC or TGW to be attached',
                   'Customer Gateway', 'Customer Gateway Address', 'Public IP address', 'ASN',
                   'Transit Gateway Attachment', 'Virtual priavte gateway']
    ws.append(column_list)

    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])
    response = client.describe_vpn_connections()
    for vpn in response['VpnConnections']:

        for tag in vpn['Tags']:
            if tag['Key'] == 'Name':
                vpn_name = tag['Value']
                vpn_type, vpn_ou, vpn_project_name, vpn_location, vpn_environment, vpn_account_name, vpn_no = vpn_name.split(
                    '-')

                attached_name = [tag['Value'] for tags in client.describe_transit_gateways(
                    Filters=[{'Name': 'transit-gateway-id', 'Values': [vpn['TransitGatewayId']]}])['TransitGateways'][0]['Tags'] if tags['Key'] == 'Name'][0]
                customer_gateway_name = ''
                customer_gateway_address = ''
                public_ip_address = ''
                asn = ''
                virtual_private_gateway = ''
                for cw in client.describe_customer_gateways(
                        Filters=[{'Name': 'customer-gateway-id', 'Values': [vpn['CustomerGatewayId']]}])['CustomerGateways']:
                    for tag in cw['Tags']:
                        if tag['Key'] == 'Name':
                            customer_gateway_name = tag['Value']
                            break
                    customer_gateway_address = cw['IpAddress']
                    asn = cw['BgpAsn']
                for ip in vpn['VgwTelemetry']:
                    public_ip_address += ip['OutsideIpAddress'] + '\n'
            ws.append([region, environment, vpn_type, vpn_ou, vpn_project_name,
                       vpn_location, vpn_environment, accountId, vpn_no, vpn_name, "", "", attached_name, customer_gateway_name, customer_gateway_address, public_ip_address, asn, "", virtual_private_gateway])


def get_tgw_data(role, accountId, region, process_type):

    global transit_gateway_header_row
    global transit_gateway_center_row
    global transit_gateway_row

    output = {"total": 0}

    ws = wb.worksheets[sheet_list.index(process_type) + 1]
    client = boto3.client('ec2', region_name=region, aws_access_key_id=role['AccessKeyId'], aws_secret_access_key=role['SecretAccessKey'],
                          aws_session_token=role['SessionToken'])

    response = client.describe_transit_gateway_route_tables()
    for tgw in response['TransitGatewayRouteTables']:
        routes = (client.search_transit_gateway_routes(
            TransitGatewayRouteTableId=tgw['TransitGatewayRouteTableId'], Filters=[{
                'Name': 'state',
                'Values': [
                        'active', 'blackhole'
                ]
            }])['Routes'])

        tgw_name = ''
        if 'Tags' in tgw:
            for tag in tgw['Tags']:
                if tag['Key'] == 'Name':
                    tgw_name = tag['Value']

        output[tgw['TransitGatewayRouteTableId']] = {
            "name": tgw_name,
            "routes": routes
        }
        output['total'] += len(routes)
    if output['total'] > 0:
        for item in output:
            if item == 'total':
                continue

            ws.merge_cells(start_row=transit_gateway_center_row, end_row=transit_gateway_center_row +
                           len(output[item]['routes'])-1, start_column=4,  end_column=4)
            ws.cell(row=transit_gateway_center_row, column=4).value = item
            ws.merge_cells(start_row=transit_gateway_center_row, end_row=transit_gateway_center_row +
                           len(output[item]['routes'])-1, start_column=5,  end_column=5)
            ws.cell(row=transit_gateway_center_row,
                    column=5).value = output[item]['name']
            transit_gateway_center_row += len(output[item]['routes'])

            for route in output[item]['routes']:
                ws.cell(
                    row=transit_gateway_row, column=6).value = route['DestinationCidrBlock']

                if 'TransitGatewayAttachments' in route:
                    ws.cell(
                        row=transit_gateway_row, column=7).value = route['TransitGatewayAttachments'][0]['TransitGatewayAttachmentId']
                    ws.cell(
                        row=transit_gateway_row, column=8).value = route['TransitGatewayAttachments'][0]['ResourceType']
                    ws.cell(
                        row=transit_gateway_row, column=9).value = route['TransitGatewayAttachments'][0]['ResourceId']

                ws.cell(
                    row=transit_gateway_row, column=10).value = route['Type']
                ws.cell(
                    row=transit_gateway_row, column=11).value = route['State']
                transit_gateway_row += 1

        ws.merge_cells(start_row=transit_gateway_header_row,
                       end_row=transit_gateway_header_row+output['total'] - 1, start_column=1,  end_column=1)
        ws.cell(row=transit_gateway_header_row,
                column=1).value = region_mapping[region]
        ws.cell(row=transit_gateway_header_row, column=1).alignment = Alignment(
            horizontal='center', vertical='center')
        ws.merge_cells(start_row=transit_gateway_header_row,
                       end_row=transit_gateway_header_row+output['total'] - 1, start_column=2,  end_column=2)
        ws.cell(row=transit_gateway_header_row, column=2).value = environment
        ws.cell(row=transit_gateway_header_row, column=2).alignment = Alignment(
            horizontal='center', vertical='center')
        ws.merge_cells(start_row=transit_gateway_header_row,
                       end_row=transit_gateway_header_row+output['total'] - 1, start_column=3,  end_column=3)
        ws.cell(row=transit_gateway_header_row, column=3).value = accountId
        ws.cell(row=transit_gateway_header_row, column=3).alignment = Alignment(
            horizontal='center', vertical='center')
        transit_gateway_header_row += output['total']


if __name__ == '__main__':
    environment = str(input('Enter the environment: '))

    with open('./setting.yaml', 'r') as f:
        scan_list = yaml.load(f, Loader=yaml.BaseLoader)
    wb = Workbook()
    ws = wb.active

    sheet_list = ['VPC', 'Route Table', 'Transit Gateway']

    def header(item):
        return {
            'VPC': ['VPC Name', 'CIDR',
                    'VPC ID', "Account ID", "Region"],
            'Route Table': ['Region', 'Environment', 'Account', 'VPC ID', 'Route Table Name', 'Route Table ID',
                            'Destination', 'Target'],
            "Transit Gateway": ['Region', 'Environment', 'Account', 'Transit Gateway route table ID', 'Transit Gateway Route Table Name', "CIDR",
                                "Transit-GW Attachment ID", "Resource type", "ID (VPC/Direct Connect Gateway/VPN)",
                                "Route type", "Route state"]
        }[item]

    for item in sheet_list:
        ws = wb.create_sheet(item)
        ws.append(header(item))

    for region in scan_list:
        for account in scan_list[region]:
            print(region, account)
            # assume role
            sts = boto3.client('sts')
            role = sts.assume_role(
                RoleArn='arn:aws:iam::{}:role/OrganizationAccountAccessRole'.format(
                    account),
                RoleSessionName='AWSCLI-Session'
            )['Credentials']
            get_vpc_data(role, account, region, 'VPC')
            # get_subnet_data()
            get_route_table_data(role, account, region, 'Route Table')
            # get_igw_data(role, account, region)
            # get_endpoint_data()
            # get_endpoint_service_data()
            # get_nat_data()
            # get_sg_data()
            # get_vpn_data(role, account, region)
            get_tgw_data(role, account, region, 'Transit Gateway')
            print('-----Finish-----')
    wb.save('output/output.xlsx')
